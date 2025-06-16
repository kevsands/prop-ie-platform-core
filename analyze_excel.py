#!/usr/bin/env python3
"""
Analyze the master property development Excel workbook
"""
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import json
import sys

def analyze_excel_workbook(file_path):
    """Analyze the Excel workbook structure and extract key information"""
    
    try:
        # Load workbook
        wb = load_workbook(file_path, data_only=True)
        
        analysis = {
            "workbook_info": {
                "total_sheets": len(wb.sheetnames),
                "sheet_names": wb.sheetnames
            },
            "sheets_analysis": {},
            "formulas": [],
            "key_metrics": {}
        }
        
        print(f"📊 Analyzing Excel Workbook: {file_path}")
        print(f"Total Sheets: {len(wb.sheetnames)}")
        print("\n🗂️  Sheet Names:")
        for i, sheet_name in enumerate(wb.sheetnames, 1):
            print(f"  {i}. {sheet_name}")
        
        # Analyze each sheet
        for sheet_name in wb.sheetnames:
            print(f"\n📋 Analyzing Sheet: {sheet_name}")
            
            ws = wb[sheet_name]
            sheet_analysis = {
                "dimensions": f"{ws.max_row} rows x {ws.max_column} columns",
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "headers": [],
                "sample_data": [],
                "formulas": [],
                "named_ranges": []
            }
            
            # Get headers (first row)
            headers = []
            for col in range(1, min(ws.max_column + 1, 21)):  # Max 20 columns for headers
                cell_value = ws.cell(row=1, column=col).value
                if cell_value:
                    headers.append(str(cell_value))
                else:
                    headers.append(f"Col_{col}")
            
            sheet_analysis["headers"] = headers
            
            # Get sample data (first 10 rows)
            sample_data = []
            for row in range(1, min(ws.max_row + 1, 11)):  # First 10 rows
                row_data = []
                for col in range(1, min(ws.max_column + 1, 21)):  # Max 20 columns
                    cell = ws.cell(row=row, column=col)
                    value = cell.value
                    
                    # Check for formulas
                    if hasattr(cell, 'data_type') and cell.data_type == 'f':
                        formula = f"={cell.value}" if cell.value else "=FORMULA"
                        sheet_analysis["formulas"].append({
                            "cell": f"{cell.coordinate}",
                            "formula": formula
                        })
                        row_data.append(f"FORMULA: {formula}")
                    else:
                        row_data.append(str(value) if value is not None else "")
                
                sample_data.append(row_data)
            
            sheet_analysis["sample_data"] = sample_data
            analysis["sheets_analysis"][sheet_name] = sheet_analysis
            
            print(f"  📐 Dimensions: {sheet_analysis['dimensions']}")
            print(f"  📝 Headers: {', '.join(headers[:10])}...")
            print(f"  🧮 Formulas found: {len(sheet_analysis['formulas'])}")
        
        # Look for named ranges
        print(f"\n🏷️  Named Ranges:")
        if hasattr(wb, 'defined_names'):
            for named_range in wb.defined_names:
                print(f"  - {named_range.name}: {named_range.attr_text}")
        
        # Try to identify key patterns
        print(f"\n🔍 Pattern Analysis:")
        
        # Look for phase-related sheets
        phase_sheets = [name for name in wb.sheetnames if 'phase' in name.lower()]
        if phase_sheets:
            print(f"  📈 Phase sheets found: {phase_sheets}")
        
        # Look for cost/revenue related sheets
        financial_sheets = [name for name in wb.sheetnames if any(keyword in name.lower() 
                           for keyword in ['cost', 'revenue', 'financial', 'budget', 'cash', 'profit'])]
        if financial_sheets:
            print(f"  💰 Financial sheets: {financial_sheets}")
        
        # Look for project-related sheets
        project_sheets = [name for name in wb.sheetnames if any(keyword in name.lower() 
                         for keyword in ['project', 'development', 'fitzgerald', 'site'])]
        if project_sheets:
            print(f"  🏗️  Project sheets: {project_sheets}")
        
        return analysis
        
    except Exception as e:
        print(f"❌ Error analyzing workbook: {e}")
        return None

def main():
    file_path = "/Users/kevin/Downloads/25,000,000.xlsx"
    
    print("🚀 Starting Excel Workbook Analysis")
    print("=" * 50)
    
    analysis = analyze_excel_workbook(file_path)
    
    if analysis:
        # Save analysis to JSON
        output_file = "/Users/kevin/backups/awsready_20250524/prop-ie-aws-app/excel_analysis.json"
        with open(output_file, 'w') as f:
            json.dump(analysis, f, indent=2, default=str)
        
        print(f"\n✅ Analysis complete! Results saved to: {output_file}")
        print("\n📊 Summary:")
        print(f"  - Total sheets: {analysis['workbook_info']['total_sheets']}")
        print(f"  - Sheet names: {', '.join(analysis['workbook_info']['sheet_names'])}")
    else:
        print("❌ Analysis failed")

if __name__ == "__main__":
    main()